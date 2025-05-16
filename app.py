from flask import Flask, render_template, request, redirect, url_for
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

ARQUIVO = 'faturamento.xlsx'

def salvar_em_planilha(novo_cadastro):
    if os.path.exists(ARQUIVO):
        wb = load_workbook(ARQUIVO)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # Cria o cabeçalho com base nas chaves
        cabecalho = list(novo_cadastro.keys())
        ws.append(cabecalho)

    # Garante que as colunas estejam na ordem do cabeçalho
    cabecalho = [col.value for col in ws[1]]
    linha = [novo_cadastro.get(col, '') for col in cabecalho]
    ws.append(linha)
    wb.save(ARQUIVO)

def carregar_cadastros():
    if not os.path.exists(ARQUIVO):
        return []
    wb = load_workbook(ARQUIVO)
    ws = wb.active
    linhas = list(ws.values)
    if not linhas or len(linhas) < 2:
        return []

    cabecalho = linhas[0]
    dados = []
    for linha in linhas[1:]:
        dados.append(dict(zip(cabecalho, linha)))
    return dados

def parse_valor(valor):
    if not valor:
        return 0.0
    valor_str = str(valor).strip()

    # Remove "R$", espaços e troca vírgula por ponto (decimal)
    valor_limpo = valor_str.replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")

    try:
        return float(valor_limpo)
    except ValueError:
        print(f"Erro ao converter valor: {valor_str}")
        return 0.0


@app.route("/")
def home():
    return render_template("index.html")

@app.route("/enviar", methods=["POST"])
def enviar():
    dados = request.form.to_dict()
    salvar_em_planilha(dados)
    return redirect(url_for('home'))
@app.route("/cadastros")
def lista_cadastros():
    cadastros = carregar_cadastros()
    
    # Somar só os valores de 'orcamento' que não estejam vazios e que possam ser convertidos para float
    total_orcamento = 0.0
    for c in cadastros:
        valor = c.get('orcamento')
        if valor:  # só se tiver algo no campo
            valor_float = parse_valor(valor)
            total_orcamento += valor_float

    return render_template("cadastros.html", cadastros=cadastros, total_orcamento=total_orcamento)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=4090, debug=True)
