from flask import Flask, render_template, request, redirect, url_for
import os
from openpyxl import Workbook, load_workbook
from flask import jsonify



app = Flask(__name__)


ARQUIVO = 'faturamento.xlsx'

def salvar_em_planilha(novo_cadastro):
    if os.path.exists(ARQUIVO):
        wb = load_workbook(ARQUIVO)
        ws = wb.active
        # Lê o cabeçalho atual
        cabecalho = [cell.value for cell in ws[1]]
        # Verifica se alguma chave nova não está no cabeçalho
        chaves_novas = [k for k in novo_cadastro.keys() if k not in cabecalho]
        if chaves_novas:
            # Adiciona colunas novas ao cabeçalho
            cabecalho.extend(chaves_novas)
            # Atualiza a primeira linha com o novo cabeçalho completo
            for col_num, valor in enumerate(cabecalho, start=1):
                ws.cell(row=1, column=col_num, value=valor)
        # Prepara a linha de dados na ordem do cabeçalho atualizado
        linha = [novo_cadastro.get(col, '') for col in cabecalho]
    else:
        wb = Workbook()
        ws = wb.active
        cabecalho = list(novo_cadastro.keys())
        ws.append(cabecalho)
        linha = [novo_cadastro.get(col, '') for col in cabecalho]

    ws.append(linha)
    wb.save(ARQUIVO)

def carregar_cadastros():
    if not os.path.exists(ARQUIVO):
        return []
    wb = load_workbook(ARQUIVO)
    ws = wb.active
    linhas = list(ws.values)
    if len(linhas) < 2:
        return []

    cabecalho = linhas[0]
    dados = [dict(zip(cabecalho, linha)) for linha in linhas[1:] if linha and any(linha)]
    return dados

def parse_valor(valor):
    if not valor:
        return 0.0
    valor_str = str(valor).strip()

    # Remove "R$", espaços e pontos como separadores de milhar
    valor_limpo = valor_str.replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")

    try:
        return float(valor_limpo)
    except ValueError:
        print(f"Erro ao converter valor: {valor_str}")
        return 0.0


@app.route("/")
def home():
    return render_template("index.html")

@app.route("/extrato")
def extrato():
    return render_template("extrato.html")



@app.route('/excluir_cadastro', methods=['POST'])
def excluir_cadastro():
    try:
        linha = int(request.form['linha']) - 1  # loop.index começa em 1
    except (ValueError, KeyError):
        return redirect(url_for('lista_cadastros'))

    cadastros = carregar_cadastros()

    if 0 <= linha < len(cadastros):
        cadastros.pop(linha)

        # Reescreve a planilha inteira com os cadastros restantes
        wb = Workbook()
        ws = wb.active

        if cadastros:
            cabecalho = list(cadastros[0].keys())
            ws.append(cabecalho)

            for cadastro in cadastros:
                linha_valores = [cadastro.get(col, '') for col in cabecalho]
                ws.append(linha_valores)

        wb.save(ARQUIVO)

    return redirect(url_for('lista_cadastros'))

@app.route("/enviar", methods=["POST"])
def enviar():
    dados = request.form.to_dict()
    salvar_em_planilha(dados)
    return redirect(url_for('home'))

@app.route("/cadastros")
def lista_cadastros():
    cadastros = carregar_cadastros()

    total_orcamento = 0.0
    total_minutos = 0  # vamos somar tudo em minutos
    projetos = set()
    categorias = set()

    for c in cadastros:
        # soma o orçamento
        valor = c.get('orcamento')
        if valor:
            total_orcamento += parse_valor(valor)

        # soma o tempo de serviço
        tempo = c.get('tempo_servico', '')
        if tempo:
            try:
                if ':' in tempo:
                    horas, minutos = map(int, tempo.split(':'))
                    total_minutos += horas * 60 + minutos
                else:
                    total_minutos += int(float(tempo) * 60)  # se for decimal
            except ValueError:
                pass  # ignora se não conseguir converter

        # Adiciona valores únicos para filtros
        if c.get("nome_projeto"):
            projetos.add(c["nome_projeto"])
        if c.get("tipo_projeto"):
            categorias.add(c["tipo_projeto"])

    total_liquido = total_orcamento / 2

    # converte minutos totais para HH:MM
    horas_totais = total_minutos // 60
    minutos_totais = total_minutos % 60
    total_horas_formatado = f"{int(horas_totais):02d}:{int(minutos_totais):02d}"

    return render_template(
        "cadastros.html",
        cadastros=cadastros,
        total_orcamento=total_orcamento,
        total_liquido=total_liquido,
        total_horas_formatado=total_horas_formatado,
        projetos=sorted(projetos),
        categorias=sorted(categorias)
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=4090, debug=True)



